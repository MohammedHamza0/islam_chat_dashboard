import openpyxl
import json
from pathlib import Path

DIR_PATH = Path(r"D:\Midade.Com\islam_chat_dashboard\LLM point extraction and analysis\Eng.Menna")

print("=== 1. Inspecting Markdown File: Analyze top conversation topics.md ===")
md_path = DIR_PATH / "Analyze top conversation topics.md"
with open(md_path, "r", encoding="utf-8", errors="ignore") as f:
    md_content = f.read()
print(f"MD file size: {len(md_content)} chars")
print("First 800 chars of MD:")
print(md_content[:800])
print("\n--- Headers / Sections in MD ---")
for line in md_content.splitlines():
    if line.startswith("#"):
        print("  ", line[:80])

def inspect_excel(excel_name):
    print(f"\n=== Inspecting Excel File: {excel_name} ===")
    wb_path = DIR_PATH / excel_name
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    print("Sheet names:", wb.sheetnames)
    for sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"\n  [Sheet: {sname}] - Dimensions: {sheet.max_row} rows x {sheet.max_column} cols")
        headers = [sheet.cell(row=1, column=col).value for col in range(1, min(sheet.max_column + 1, 15))]
        print(f"    Headers: {headers}")
        if sheet.max_row > 1:
            first_row = [sheet.cell(row=2, column=col).value for col in range(1, min(sheet.max_column + 1, 15))]
            print(f"    Sample Row 1: {first_row[:8]}")

inspect_excel("existing_muslim_dawah_bot_analysis.xlsx")
inspect_excel("islamic_dawah_topics_analysis.xlsx")
inspect_excel("religion_specific_concern_playbooks.xlsx")

print("\n=== 5. Inspecting HTML: islamic_dawah_bot_interactive_dashboard_ar.html ===")
html_path = DIR_PATH / "islamic_dawah_bot_interactive_dashboard_ar.html"
with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
    html_c = f.read()
print(f"HTML size: {len(html_c)} chars, Title preview:")
for line in html_c.splitlines()[:20]:
    if "<title>" in line or "<h1>" in line or "<h2>" in line:
        print("  ", line.strip())
